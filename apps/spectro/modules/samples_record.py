"""
apps/spectro/modules/samples_record.py

Module for the "Samples Record".
"""
import io, re

import openpyxl
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from django.contrib.auth.decorators import login_required
from django.http import HttpResponse, JsonResponse
from django.shortcuts import render
from django.urls import reverse
from django.utils import timezone

from apps.spectro.models import (
    SpectrometerRecord,
    SpectroStandard,
    StdLimitChangelog,
    LotSample,
    SpectroJudgement,
    VisualJudgement,
    SpecialCase,
    SpecialCaseChangelog,
    QcProgramRecord,
)

def format_datetime_no_leading_zeros(dt):
    if not dt:
        return "-"
    return "{}/{}/{} {}:{:02d}".format(dt.month, dt.day, dt.year, dt.hour, dt.minute)

# ---------------------------------------------------------------------
# QC lot-format parsing (Task 6.2)
#
# QC's sticker_lot / internal_lot columns are free text, not a fixed
# single-lot format like this project's own LotSample.sample_name.
# The SAME column can encode any of:
#   (a)/(b)  a start-end RANGE of lot codes    "7382AO-7385AO"
#   (c)/(d)  a lot code + embedded BAG number   "7329AL(16)" / "7400AL(1-21)"
#   (e)      a lot code + embedded INTERNAL LOT alias  "5853AL(5873AL)"
#   (f)/(g)  a lot code + space-separated BAG   "7400AL 12" / "7400AL 12-13"
#   plain    just the lot code on its own       "7383AO"
# ---------------------------------------------------------------------

LOT_CODE_RE = re.compile(r'^(\d{1,6})([A-Za-z]{1,3})$')

QC_RANGE_RE = re.compile(r'^(\d{1,6}[A-Za-z]{1,3})\s*-\s*(\d{1,6}[A-Za-z]{1,3})$')
QC_PAREN_RE = re.compile(r'^(\d{1,6}[A-Za-z]{1,3})\s*\(\s*([^)]+?)\s*\)$')
QC_SPACE_RE = re.compile(r'^(\d{1,6}[A-Za-z]{1,3})\s+(\d+(?:-\d+)?)$')
QC_PLAIN_RE = re.compile(r'^(\d{1,6}[A-Za-z]{1,3})$')
# a paren's inner content counts as "another lot code" (format e) only
# if it has letters in it -- a pure digit/digit-range inner value is a
# bag number (formats c/d) instead
QC_LOT_LIKE_RE = re.compile(r'^\d{1,6}[A-Za-z]{1,3}$')


def _split_lot_code(code):
    """'7382AO' -> (7382, 'AO'). None if it doesn't match the expected shape."""
    m = LOT_CODE_RE.match(code.strip().upper())
    if not m:
        return None
    return int(m.group(1)), m.group(2)


def _classify_qc_lot_text(text):
    """
    Classifies a single QC sticker_lot/internal_lot text value into one
    of the known (a)-(g) shapes, WITHOUT yet comparing it against any
    particular (h) lot. Kept separate from matching so callers can
    distinguish "this text is a genuinely recognized shape that just
    doesn't include our lot" from "this text doesn't match any known
    QC lot-format convention at all" -- the latter is a QC data-quality
    signal, not a plain not-found.

    Returns None for a blank/empty field, otherwise a dict shaped:
        {"kind": "range", "start": (num, suffix)|None, "end": (num, suffix)|None, "raw": text}
        {"kind": "paren", "base": (num, suffix)|None, "inner": <str>, "raw": text}
        {"kind": "space", "base": (num, suffix)|None, "bag": <str>, "raw": text}
        {"kind": "plain", "base": (num, suffix)|None, "raw": text}
        {"kind": "unrecognized", "raw": text}
    """
    if not text:
        return None
    raw = text.strip()
    if not raw:
        return None

    m = QC_RANGE_RE.match(raw)
    if m:
        return {
            "kind": "range",
            "start": _split_lot_code(m.group(1)),
            "end": _split_lot_code(m.group(2)),
            "raw": raw,
        }

    m = QC_PAREN_RE.match(raw)
    if m:
        return {
            "kind": "paren",
            "base": _split_lot_code(m.group(1)),
            "inner": m.group(2).strip().upper(),
            "raw": raw,
        }

    m = QC_SPACE_RE.match(raw)
    if m:
        return {
            "kind": "space",
            "base": _split_lot_code(m.group(1)),
            "bag": m.group(2).strip(),
            "raw": raw,
        }

    m = QC_PLAIN_RE.match(raw)
    if m:
        return {"kind": "plain", "base": _split_lot_code(m.group(1)), "raw": raw}

    return {"kind": "unrecognized", "raw": raw}


def _match_qc_classification(h_split, classification):
    """
    Given a parsed QC text shape (see _classify_qc_lot_text) and (h)
    split into (number, suffix), attempts an actual match.

    Returns a (outcome, detail) tuple:
        ("match", {"bag": <str>|None, "internal_lot": <str>|None})
        ("suffix_anomaly", <raw text>)  -- range's own start/end suffix
                                            letters disagree with each
                                            other (QC data-quality issue,
                                            not a normal no-match)
        ("no_match", None)
    """
    kind = classification["kind"]

    if kind == "range":
        start, end = classification["start"], classification["end"]
        if not start or not end:
            return ("no_match", None)
        if start[1] != end[1]:
            return ("suffix_anomaly", classification["raw"])
        if not h_split or h_split[1] != start[1]:
            return ("no_match", None)
        lo, hi = sorted((start[0], end[0]))
        if lo <= h_split[0] <= hi:
            return ("match", {"bag": None, "internal_lot": None})
        return ("no_match", None)

    if kind == "paren":
        base = classification["base"]
        if not base or not h_split or base != h_split:
            return ("no_match", None)
        inner = classification["inner"]
        if QC_LOT_LIKE_RE.match(inner):
            return ("match", {"bag": None, "internal_lot": inner})   # (e)
        return ("match", {"bag": inner, "internal_lot": None})        # (c)/(d)

    if kind == "space":
        base = classification["base"]
        if not base or not h_split or base != h_split:
            return ("no_match", None)
        return ("match", {"bag": classification["bag"], "internal_lot": None})  # (f)/(g)

    if kind == "plain":
        base = classification["base"]
        if base and h_split and base == h_split:
            return ("match", {"bag": None, "internal_lot": None})
        return ("no_match", None)

    # "unrecognized" -- caller tracks this as its own anomaly bucket
    return ("no_match", None)


def _qc_lookup_for_row(product_code, sample_name):
    """
    Task 6: cross-checks a LotSample's product code + lot number
    (sample_name) against QcProgramRecord -- the read-only view_spectro
    view on the 'server' database (see apps/core/db_router.py).

    Step order matches the spec exactly:
      0. sample_name carries the "LT " or "DR " reference-reading
         prefix (see LOT_NUMBER_RE / samples_reader.py)?  yes ->
         reference, stop here -- these are calibration-adjacent
         readings, never expected to exist in QC at all.
      1. product_code present in QC?  no -> warning, stop here.
      2. sample_name matches sticker_lot OR internal_lot for that
         product code?  no -> warning, stop here.
         ("only allow to return one true" -- first match wins; a lot
         number legitimately appearing in QC should only match one of
         the two fields for a given product code.)
      3/4. both true -> pull bag_number and status off that QC row,
         "N/A" each if blank.

    Returns:
        {
            "match": "ok" | "warning" | "anomaly" | "reference",
            "message": <user-facing string>,
            "bag": <str>,            # bag_number or "N/A"
            "status_value": <str>|None,  # QC's own "status" column, only set when match == "ok"
        }
    """
    name_check = (sample_name or "").strip().upper()
    if name_check.startswith("LT ") or name_check.startswith("DR "):
        return {
            "match": "reference",
            "message": "This is a reference lot and not expected to exist in QC.",
            "bag": "N/A",
            "internal_lot": "-",
            "status_value": None,
        }

    if not product_code:
        return {
            "match": "warning",
            "message": "Selected product code has no record in QC.",
            "bag": "N/A",
            "internal_lot": "-",
            "status_value": None,
        }

    # (1) product codes are always uppercase by this project's own
    # PRODUCT_CODE_RE convention -- but QC's own data isn't guaranteed
    # to be, so filter case-insensitively rather than assume QC's
    # casing matches. Display values are separately normalized to caps
    # further down.
    qc_rows_for_product = list(
        QcProgramRecord.objects.filter(product_code__iexact=product_code).order_by("qc_id")
    )
    if not qc_rows_for_product:
        return {
            "match": "warning",
            "message": "Selected product code has no record in QC.",
            "bag": "N/A",
            "internal_lot": "-",
            "status_value": None,
        }

    name = (sample_name or "").strip().upper()
    h_split = _split_lot_code(name)

    # (0) 3rd/final data-cleaning layer: Spectro's own LOT_NUMBER_RE
    # (samples_reader.py, enforced client- and server-side) should make
    # this unreachable through the app itself -- but a lot inserted
    # directly via DB/admin could still bypass that. Without this check
    # such a value would silently fall through every branch below to
    # "no_match" and surface as a generic red "not found in QC", which
    # is misleading -- the real problem is Spectro's own data, not QC's.
    if name and h_split is None:
        return {
            "match": "anomaly",
            "message": "Typographical spectro lot number",
            "bag": "N/A",
            "internal_lot": "-",
            "status_value": None,
        }

    matched_row = None
    matched_result = None
    suffix_anomaly_raw = None   # (2) range whose own start/end suffix letters disagree
    unrecognized_seen = False   # (3) QC text that doesn't fit any known (a)-(g) shape

    for qc_row in qc_rows_for_product:
        for field_text in (qc_row.sticker_lot, qc_row.internal_lot):
            classification = _classify_qc_lot_text(field_text)
            if classification is None:
                continue

            if classification["kind"] == "unrecognized":
                unrecognized_seen = True
                continue

            outcome, detail = _match_qc_classification(h_split, classification)

            if outcome == "suffix_anomaly":
                if suffix_anomaly_raw is None:
                    suffix_anomaly_raw = detail
                continue

            if outcome == "match":
                matched_row = qc_row
                matched_result = detail
                break

        if matched_row is not None:
            break  # (4) "only allow to return one true" -- first match wins, stop scanning

    if matched_row is not None:
        # (1) bag_number DB column wins if populated; fall back to
        # whatever was parsed out of the sticker/internal lot TEXT
        # (formats c/d/f/g) when that column is empty. Always shown
        # in caps regardless of how QC actually stored it.
        if matched_row.bag_number:
            bag = matched_row.bag_number.strip().upper()
        elif matched_result["bag"]:
            bag = matched_result["bag"].strip().upper()
        else:
            bag = "N/A"

        # (5) format (e) is currently the ONLY source for Internal Lot.
        internal_lot = (
            matched_result["internal_lot"].strip().upper()
            if matched_result["internal_lot"] else "-"
        )

        return {
            "match": "ok",
            "message": "This Lot/Bag is found in QC.",
            "bag": bag,
            "internal_lot": internal_lot,
            "status_value": matched_row.status if matched_row.status else "N/A",
        }

    # (2) QC's own range data is internally inconsistent -- flag as a
    # QC data-quality anomaly, distinct from a plain not-found.
    if suffix_anomaly_raw:
        return {
            "match": "anomaly",
            "message": (
                'QC has a lot range for this product code with mismatched '
                'suffix letters ("' + suffix_anomaly_raw + '") -- flagging '
                'for manual review, could not confirm a match.'
            ),
            "bag": "N/A",
            "internal_lot": "-",
            "status_value": None,
        }

    # (3) QC has data for this product code, and at least one lot field
    # exists, but it doesn't fit any recognized (a)-(g) shape -- tell
    # the user data exists but is unreadable, not that it's empty.
    if unrecognized_seen:
        return {
            "match": "anomaly",
            "message": (
                "QC has a record for this product code, but its lot "
                "format could not be read — please verify manually."
            ),
            "bag": "N/A",
            "internal_lot": "-",
            "status_value": None,
        }

    return {
        "match": "warning",
        "message": "This lot is not found in QC.",
        "bag": "N/A",
        "internal_lot": "-",
        "status_value": None,
    }


def _serialize_lot_sample_row(lot_sample, standards_id, product_code=None):
    raw = lot_sample.raw_values.order_by("-date_time").first()
    delta = raw.delta_values.order_by("-date_time").first() if raw else None
    spectro_j = (
        SpectroJudgement.objects
        .filter(lot_sample=lot_sample, standard_id=standards_id)
        .order_by("-date_time")
        .first()
    )
    visual_j = VisualJudgement.objects.filter(lot_sample=lot_sample).order_by("-date_time").first()
    special = SpecialCase.objects.filter(lot_sample=lot_sample).order_by("-date_time").first()

    def num(value):
        return float(value) if value is not None else None

    if spectro_j is None:
        spectro_judgement = "PASSED"  # default is_pass = True when no record exists yet
    elif spectro_j.is_pass is None:
        spectro_judgement = "PASSED"  # is_pass defaults to True
    else:
        spectro_judgement = "PASSED" if spectro_j.is_pass else "FAILED"

    if visual_j is None or visual_j.is_pass is None:
        visual_judgement = ""
    else:
        visual_judgement = "Pass" if visual_j.is_pass else "Fail"

    qc_info = _qc_lookup_for_row(product_code, lot_sample.sample_name)

    return {
        "id": lot_sample.lot_samples_id,
        "lotSampleId": lot_sample.lot_samples_id,
        "visualJudgementId": visual_j.visual_status_id if visual_j else None,
        "spectroJudgementId": spectro_j.spectro_status_id if spectro_j else None,
        "colorSimulation": lot_sample.color_simulation or "-",
        "dateTime": format_datetime_no_leading_zeros(lot_sample.date_time),
        "stickerLot": lot_sample.sample_name,
        "bag": lot_sample.bag if lot_sample.bag else "N/A",
        "internalLot": qc_info["internal_lot"],
        "de00": num(delta.delta_e) if delta else None,
        "L": num(raw.raw_l) if raw else None,
        "C": num(raw.raw_c) if raw else None,
        "h": num(raw.raw_h) if raw else None,
        "a": num(raw.raw_a) if raw else None,
        "b": num(raw.raw_b) if raw else None,
        "dL": num(delta.delta_l) if delta else None,
        "dC": num(delta.delta_c) if delta else None,
        "dH": num(delta.delta_h) if delta else None,
        "da": num(delta.delta_a) if delta else None,
        "db": num(delta.delta_b) if delta else None,
        "colorOffset": spectro_j.color_offset if spectro_j and spectro_j.color_offset else "-",
        "spectroJudgement": spectro_judgement,
        "visualJudgement": visual_judgement,
        "finalQcEval": qc_info["status_value"] if qc_info["match"] == "ok" else "-",
        "qcMatch": qc_info["match"],      # "ok" | "warning" -- drives column 2's icon
        "qcMessage": qc_info["message"],  # tooltip text for column 2
        "reasonIfFail": visual_j.visual_fail_reason if visual_j and visual_j.visual_fail_reason else "",
        "spectroRemarks": spectro_j.spectro_remarks if spectro_j and spectro_j.spectro_remarks else "",
        "specialPass": bool(special.is_pass) if special else False,
        "specialPassBy": special.passed_by if special and special.passed_by else "",
    }

def _recalculate_spectro_judgements(record, threshold):
    """
    Re-evaluate is_pass for every lot sample under this record's standards
    against the new threshold. Matches by lot_sample (+ its standard) and
    updates the existing SpectroJudgement row in place -- same pattern as
    save_visual_judgement -- instead of inserting a new row per change.
    """
    standards = SpectroStandard.objects.filter(record=record)
    lot_samples = LotSample.objects.filter(standard__in=standards)

    for lot_sample in lot_samples:
        raw = lot_sample.raw_values.order_by("-date_time").first() # type: ignore
        delta = raw.delta_values.order_by("-date_time").first() if raw else None
        if delta is None:
            continue

        is_pass = float(delta.delta_e) <= threshold

        spectro_j, _ = SpectroJudgement.objects.get_or_create(
            lot_sample=lot_sample,
            defaults={"is_pass": is_pass, "standard": lot_sample.standard},
        )
        spectro_j.is_pass = is_pass
        spectro_j.standard = lot_sample.standard
        spectro_j.save(update_fields=["is_pass", "standard"])


@login_required
def render_samples_record(request):
    context = {
        "search_product_codes_url": reverse("api_search_product_codes"),
    }
    return render(request, "pages/samples_record.html", context)


@login_required
def search_product_codes(request):
    """
    Task 1: debounced (client-side), server-side product code lookup for
    Samples Record's combobox. Replaces dumping every product code to the
    client at page load -- only codes matching the current query are ever
    sent, and only once the user has actually typed something.
    """
    query = request.GET.get("q", "").strip()
    if not query:
        return JsonResponse({"results": []})

    codes = list(
        SpectrometerRecord.objects
        .filter(product_code__icontains=query)
        .order_by("product_code")
        .values_list("product_code", flat=True)
        .distinct()[:20]
    )
    return JsonResponse({"results": codes})


@login_required
def get_lot_samples_for_standard(request):
    standards_id = request.GET.get("standards_id", "").strip()
    if not standards_id:
        return JsonResponse({"tone": "danger", "message": "standards_id is required."}, status=400)

    standard = SpectroStandard.objects.filter(pk=standards_id).select_related("record").first()
    if not standard:
        return JsonResponse({"tone": "danger", "message": "Standard not found."}, status=404)

    lot_samples = (
        LotSample.objects
        .filter(standard_id=standards_id)
        .order_by("-date_time", "-lot_samples_id")
    )

    product_code = standard.record.product_code if standard.record else None
    rows = [_serialize_lot_sample_row(ls, standards_id, product_code) for ls in lot_samples]

    return JsonResponse({"tone": "success", "message": "Samples loaded.", "rows": rows})


@login_required
def save_visual_judgement(request):
    if request.method != "POST":
        return JsonResponse({"tone": "danger", "message": "That's a bad way to check a program, brotha."}, status=405)

    lot_sample_id = request.POST.get("lot_sample_id", "").strip()
    value = request.POST.get("value", "").strip()  # "Pass" | "Fail" | ""

    if not lot_sample_id:
        return JsonResponse({"tone": "danger", "message": "lot_sample_id is required."}, status=400)

    lot_sample = LotSample.objects.filter(pk=lot_sample_id).first()
    if not lot_sample:
        return JsonResponse({"tone": "danger", "message": "Sample not found."}, status=404)

    is_pass = {"Pass": True, "Fail": False}.get(value, None)

    visual_j, _ = VisualJudgement.objects.get_or_create(
        lot_sample=lot_sample,
        defaults={"is_pass": is_pass, "judged_by": request.user.get_full_name() or request.user.username},
    )
    visual_j.is_pass = is_pass
    visual_j.judged_by = request.user.get_full_name() or request.user.username
    visual_j.save(update_fields=["is_pass", "judged_by"])

    return JsonResponse({"tone": "success", "message": "Visual judgement saved."})


@login_required
def save_visual_fail_reason(request):
    if request.method != "POST":
        return JsonResponse({"tone": "danger", "message": "That's a bad way to check a program, brotha."}, status=405)

    lot_sample_id = request.POST.get("lot_sample_id", "").strip()
    reason = request.POST.get("reason", "").strip()

    lot_sample = LotSample.objects.filter(pk=lot_sample_id).first()
    if not lot_sample:
        return JsonResponse({"tone": "danger", "message": "Sample not found."}, status=404)

    visual_j, _ = VisualJudgement.objects.get_or_create(
        lot_sample=lot_sample,
        defaults={"judged_by": request.user.get_full_name() or request.user.username},
    )
    visual_j.visual_fail_reason = reason
    visual_j.save(update_fields=["visual_fail_reason"])

    return JsonResponse({"tone": "success", "message": "Remarks saved."})


@login_required
def save_spectro_remarks(request):
    if request.method != "POST":
        return JsonResponse({"tone": "danger", "message": "That's a bad way to check a program, brotha."}, status=405)

    lot_sample_id = request.POST.get("lot_sample_id", "").strip()
    standards_id = request.POST.get("standards_id", "").strip()
    remarks = request.POST.get("remarks", "").strip()

    lot_sample = LotSample.objects.filter(pk=lot_sample_id).first()
    if not lot_sample or not standards_id:
        return JsonResponse({"tone": "danger", "message": "Sample or standard not found."}, status=404)

    spectro_j, _ = SpectroJudgement.objects.get_or_create(
        lot_sample=lot_sample,
        defaults={"spectro_remarks": remarks, "standard_id": standards_id},
    )
    spectro_j.spectro_remarks = remarks
    spectro_j.standard_id = standards_id # type: ignore
    spectro_j.save(update_fields=["spectro_remarks", "standard_id"])

    return JsonResponse({"tone": "success", "message": "Remarks saved."})


@login_required
def save_special_pass_by(request):
    if request.method != "POST":
        return JsonResponse({"tone": "danger", "message": "That's a bad way to check a program, brotha."}, status=405)

    lot_sample_id = request.POST.get("lot_sample_id", "").strip()
    value = request.POST.get("value", "").strip()

    if not lot_sample_id:
        return JsonResponse({"tone": "danger", "message": "lot_sample_id is required."}, status=400)
    if not value:
        return JsonResponse({"tone": "danger", "message": "Please select who passed this sample."}, status=400)

    lot_sample = LotSample.objects.filter(pk=lot_sample_id).first()
    if not lot_sample:
        return JsonResponse({"tone": "danger", "message": "Sample not found."}, status=404)

    special_case = SpecialCase.objects.filter(lot_sample=lot_sample).order_by("-date_time").first()

    if special_case and special_case.passed_by and special_case.passed_by != value:
        SpecialCaseChangelog.objects.create(
            special_case_ref=special_case,
            is_pass=special_case.is_pass,
            old_passed_by=special_case.passed_by,
            user=request.user,
        )

    if special_case:
        special_case.is_pass = True
        special_case.passed_by = value
        special_case.save(update_fields=["is_pass", "passed_by"])
    else:
        special_case = SpecialCase.objects.create(
            lot_sample=lot_sample,
            is_pass=True,
            passed_by=value,
        )

    return JsonResponse({"tone": "success", "message": "Special pass saved.", "passed_by": value})


@login_required
def save_std_delta_e_used(request):
    if request.method != "POST":
        return JsonResponse({"tone": "danger", "message": "That's a bad way to check a program, brotha."}, status=405)

    product_code = request.POST.get("product_code", "").strip()
    new_value_raw = request.POST.get("new_value", "").strip()

    if not product_code:
        return JsonResponse({"tone": "danger", "message": "product_code is required."}, status=400)

    if not new_value_raw:
        return JsonResponse({"tone": "danger", "message": "Standard ΔE value is required."}, status=400)

    try:
        new_value = float(new_value_raw)
    except ValueError:
        return JsonResponse({"tone": "danger", "message": "Standard ΔE must be a valid number."}, status=400)

    record = SpectrometerRecord.objects.filter(product_code=product_code).first()
    if not record:
        return JsonResponse({"tone": "danger", "message": "No spectrometer record found for this product code."}, status=404)

    old_value = record.std_delta_e_used

    if old_value is not None and new_value <= float(old_value):
        return JsonResponse({
            "tone": "danger",
            "message": "New value must be greater than the current Standard ΔE Used.",
        }, status=400)

    StdLimitChangelog.objects.create(
        record=record,
        old_std_delta_e=old_value,
        changed_by=request.user.get_full_name() or request.user.username,
        user=request.user,
    )

    record.std_delta_e_used = new_value # type: ignore
    record.save(update_fields=["std_delta_e_used"])

    _recalculate_spectro_judgements(record, new_value)

    return JsonResponse({
        "tone": "success",
        "message": "Standard ΔE Used updated successfully.",
        "std_delta_e_used": new_value,
    })


REPORT_COLUMNS = [
    ("colorSimulation", "Color Simulation"),
    ("dateTime", "Date Time"),
    ("code", "Code"),
    ("stickerLot", "Sticker Lot Number"),
    ("bag", "Bag Number"),
    ("internalLot", "Internal Lot"),
    ("de00", "ΔE*00"),
    ("L", "L*"),
    ("C", "C*"),
    ("h", "h°"),
    ("a", "a*"),
    ("b", "b*"),
    ("dL", "ΔL*"),
    ("dC", "ΔC*"),
    ("dH", "ΔH*"),
    ("da", "Δa*"),
    ("db", "Δb*"),
    ("colorOffset", "Color Offset"),
    ("spectroJudgement", "Spectro Judgement"),
    ("stdDeUsed", "STD ΔE used"),
    ("visualJudgement", "Visual Judgement"),
    ("finalQcEval", "Final QC Evaluation"),
    ("reasonIfFail", "Reason for Fail (if not color)"),
    ("spectroRemarks", "Spectro Remarks"),
    ("specialPass", "Special Pass?"),
    ("specialPassBy", "Special Pass BY:"),
]

REPORT_COLUMN_WIDTHS = [20, 17, 16, 22, 12, 12, 9, 9, 9, 9, 9, 9, 9, 9, 9, 9, 9, 14, 16, 13, 16, 18, 24, 20, 12, 16]


@login_required
def export_samples_report(request):
    """
    Builds the "Generate Report for this Product Code" Excel export.
    Pulls straight from the database using standards_id -- same query
    used to populate the table -- so the export always matches what the
    currently selected product code and standard are showing on screen.
    """
    standards_id = request.GET.get("standards_id", "").strip()
    if not standards_id:
        return JsonResponse({"tone": "danger", "message": "standards_id is required."}, status=400)

    standard = SpectroStandard.objects.filter(pk=standards_id).select_related("record").first()
    if not standard:
        return JsonResponse({"tone": "danger", "message": "Standard not found."}, status=404)

    record = standard.record
    product_code = record.product_code
    std_de_used = float(record.std_delta_e_used) if record.std_delta_e_used is not None else None

    lot_samples = LotSample.objects.filter(standard_id=standards_id).order_by("-date_time", "-lot_samples_id")
    rows = [_serialize_lot_sample_row(ls, standards_id, product_code) for ls in lot_samples]
    # export needs real datetime objects (not the UI's display string) so Excel can sort/filter dates.
    # Excel/openpyxl can't hold timezone-aware datetimes, so convert to local time and strip tzinfo.
    for lot_sample, row in zip(lot_samples, rows):
        dt = lot_sample.date_time
        if dt and timezone.is_aware(dt):
            dt = timezone.localtime(dt).replace(tzinfo=None)
        row["dateTime"] = dt
        row["code"] = product_code
        row["stdDeUsed"] = std_de_used

    # Task 5: DR-prefixed sticker lots first, then LT-prefixed, then
    # everything else -- same ordering as the Samples Record table's
    # default sort. Python's sort is stable, so rows within each group
    # keep their original (-date_time, -lot_samples_id) order.
    def _sort_rank(row):
        name = (row.get("stickerLot") or "").strip().upper()
        if name.startswith("DR"):
            return 0
        if name.startswith("LT"):
            return 1
        return 2

    rows.sort(key=_sort_rank)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"{product_code} from_spectro"[:31]

    header_font = Font(name="Roboto", size=11, bold=True, color="FF000000")
    header_fill = PatternFill(fill_type="solid", fgColor="FFFFFF00")

    for idx, (_key, label) in enumerate(REPORT_COLUMNS, start=1):
        cell = ws.cell(row=1, column=idx, value=label)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="left")

    for r_idx, row in enumerate(rows, start=2):
        for c_idx, (key, _label) in enumerate(REPORT_COLUMNS, start=1):
            value = row.get(key)
            cell = ws.cell(row=r_idx, column=c_idx, value=value)

            if key == "dateTime" and value:
                cell.number_format = "m/d/yyyy h:mm"

            if key == "colorSimulation" and isinstance(value, str) and value.startswith("#") and len(value) in (7, 9):
                hexval = value.lstrip("#")
                if len(hexval) == 6:
                    # plain RGB, no alpha in the source string
                    hexval = "FF" + hexval
                elif len(hexval) == 8:
                    # source string is stored/rendered by the browser as
                    # #RRGGBBAA (CSS Color 4 8-digit hex order), NOT
                    # #AARRGGBB. openpyxl's fgColor expects true ARGB, so
                    # reorder here rather than passing the raw channels
                    # straight through -- otherwise Excel's fill color
                    # doesn't match what's shown in the Samples Record
                    # table (e.g. pink on screen, blue in the export).
                    rr, gg, bb, aa = hexval[0:2], hexval[2:4], hexval[4:6], hexval[6:8]
                    hexval = aa + rr + gg + bb
                try:
                    cell.fill = PatternFill(fill_type="solid", fgColor=hexval)
                except ValueError:
                    pass

    last_row = max(len(rows) + 1, 2)
    last_col_letter = get_column_letter(len(REPORT_COLUMNS))

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{last_col_letter}{last_row}"

    green = Font(color="FF00B050")
    red = Font(color="FFFF0000")
    for rng in (f"S2:S{last_row}", f"U2:U{last_row}"):
        ws.conditional_formatting.add(rng, CellIsRule(operator="equal", formula=['"Pass"'], font=green))
        ws.conditional_formatting.add(rng, CellIsRule(operator="equal", formula=['"Passed"'], font=green))
        ws.conditional_formatting.add(rng, CellIsRule(operator="equal", formula=['"Fail"'], font=red))
        ws.conditional_formatting.add(rng, CellIsRule(operator="equal", formula=['"Failed"'], font=red))

    for idx, w in enumerate(REPORT_COLUMN_WIDTHS, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = w

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    filename = f"{product_code} from_spectro.xlsx"
    response = HttpResponse(
        buffer.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response


@login_required
def get_standards_for_product_code(request):
    product_code = request.GET.get("product_code", "").strip()
    if not product_code:
        return JsonResponse({"tone": "danger", "message": "product_code is required."}, status=400)

    record = SpectrometerRecord.objects.filter(product_code=product_code).first()
    if not record:
        return JsonResponse({
            "tone": "info",
            "message": "No spectrometer record found for this product code.",
            "standards": [],
            "std_delta_e_used": None,
        })

    standards = list(
        SpectroStandard.objects
        .filter(record=record)
        .order_by("-is_active_standard", "-date_time")
        .values("standards_id", "standard_name", "raw_l", "raw_a", "raw_b")
    )

    return JsonResponse({
        "tone": "success",
        "message": "Standards loaded.",
        "standards": standards,
        "std_delta_e_used": (
            float(record.std_delta_e_used) if record.std_delta_e_used is not None else None
        ),
    })

